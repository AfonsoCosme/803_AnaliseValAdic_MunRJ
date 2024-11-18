# View.py
import openpyxl
from openpyxl.styles import PatternFill, Font, numbers, NamedStyle, Alignment
import pandas as pd
from typing import Dict
import logging
import time
import locale
import configparser
import json
from pathlib import Path

class ExcelView:
    def __init__(self, config: configparser.ConfigParser, project_root: Path):
        """Inicialização da classe ExcelView com configurações de formatação."""
        self.workbook = None
        self.config = config
        self.project_root = project_root
        
        # Configurações de log
        logging.basicConfig(filename='excel_view.log', level=logging.INFO)
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        
        # Carregamento de configurações de formatação
        self.start_row = self.config.getint('FORMATTING', 'start_row')
        self.report_font = self.config['FORMATTING']['font_normal']
        self.report_font_size = self.config.getint('FORMATTING', 'font_size_normal')
        self.accounting_format = self.config['FORMATTING']['accounting_format']
        self.text_alignment = self.config['FORMATTING']['text_align_left']

    def setup_accounting_style(self, workbook):
        """Configura o estilo contábil para o workbook."""
        if "accounting_style" not in workbook.named_styles:
            accounting_style = NamedStyle(name="accounting_style")
            accounting_style.number_format = self.accounting_format
            workbook.add_named_style(accounting_style)

    def _update_sheet(self, sheet, df: pd.DataFrame) -> None:
        """Atualiza a planilha com os dados do DataFrame, aplicando formatação."""
        configured_font = Font(name=self.report_font, size=self.report_font_size)

        # Colunas com formato GERAL
        general_format_columns = [
            'SigMun', 
            'MUNICIPIO', 
            'CPF_CNPJ', 
            'RazSoc', 
            'ANO', 
            'InscEst'
        ]

        # Cor de fundo para valores negativos (vermelho claro)
        light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Aplicação de formatação por coluna
        for r_idx, row in enumerate(df.itertuples(index=False), start=self.start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.font = configured_font

                # Determina o tipo de coluna para formatação
                header_cell = sheet.cell(row=self.start_row - 1, column=c_idx)
                column_name = header_cell.value

                # Formatação para colunas de texto (Formato GERAL)
                if column_name in general_format_columns:
                    cell.number_format = 'General'
                    cell.alignment = Alignment(horizontal='left')
                
                # Formatação para colunas numéricas
                else:
                    cell.number_format = self.accounting_format
                    cell.alignment = Alignment(horizontal='right')
                    
                    # Aplica preenchimento vermelho claro para valores negativos
                    if isinstance(value, (int, float)) and value < 0:
                        cell.fill = light_red_fill

    def update_tab_unificada(self, df: pd.DataFrame) -> None:
        """Atualiza a aba TAB_Unificada."""
        sheet = self.workbook['TAB_Unificada']
        df = df[['SigMun', 'MUNICIPIO', 'InscEst', 'CPF_CNPJ', 'RazSoc', 'ANO', 'VALOR']]
        self._update_sheet(sheet, df)
        logging.info(f"Aba TAB_Unificada atualizada. Linhas processadas: {len(df)}")

    def update_tab_evolrazsoc(self, df: pd.DataFrame) -> None:
        """Atualiza a aba TAB_EvolRazSoc."""
        sheet = self.workbook['TAB_EvolRazSoc']
        self._update_sheet(sheet, df)
        logging.info(f"Aba TAB_EvolRazSoc atualizada. Linhas processadas: {len(df)}")

    def update_analysis_tabs(self, df_analysis: Dict[str, pd.DataFrame]) -> None:
        """Atualiza as abas de análise para cada município."""
        for sig_mun, df in df_analysis.items():
            sheet_name = f"Variacao{sig_mun}"
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                self._update_sheet(sheet, df)
                logging.info(f"Aba {sheet_name} atualizada. Linhas processadas: {len(df)}")
            else:
                logging.warning(f"Aba {sheet_name} não encontrada no arquivo Excel.")

    def update_excel(self, file_path: str, df_unified: pd.DataFrame, df_evol: pd.DataFrame, df_analysis: Dict[str, pd.DataFrame]) -> None:
        """Método principal para atualizar o arquivo Excel."""
        start_time = time.time()
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            
            # Configura o estilo contábil
            self.setup_accounting_style(self.workbook)
            
            self.update_tab_unificada(df_unified)
            self.update_tab_evolrazsoc(df_evol)
            self.update_analysis_tabs(df_analysis)
            
            self.workbook.save(file_path)
            end_time = time.time()
            logging.info(f"Arquivo Excel atualizado: {file_path}. Tempo total: {end_time - start_time:.2f} segundos")
        except Exception as e:
            logging.error(f"Erro ao atualizar o arquivo Excel: {str(e)}")
            raise
        finally:
            # Garantir que o workbook seja fechado mesmo se ocorrer um erro
            if hasattr(self, 'workbook') and self.workbook is not None:
                self.workbook.close()
                del self.workbook