### DataAnalyzer.py
import os
import json
import logging
from datetime import datetime
import configparser
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.worksheet.page import PageMargins
from typing import Dict, Tuple

class DataAnalyzer:
    def __init__(self, projectRoot: str, config: configparser.ConfigParser):
        """
        Initialize the DataAnalyzer with the project root directory and configuration settings.

        :param projectRoot: Root directory of the project
        :param config: ConfigParser object containing configuration settings
        """
        self.projectRoot = projectRoot
        self.config = config
        self.sigMunMap = {"Areal": "ARE", "Itaguai": "ITG", "Porto Real": "POR"}
        self.formatKeywords = self.loadFormatKeywords()
        self.setupLogging()

    def setupLogging(self):
        """Configure the logging system to record debug and information messages."""
        logPath = os.path.join(self.projectRoot, "logs", "data_analyzer.log")
        os.makedirs(os.path.dirname(logPath), exist_ok=True)

        logging.basicConfig(
            filename=logPath,
            level=logging.INFO,  # INFO, DEBUG, WARNING
            format='%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

    def loadFormatKeywords(self):
        """
        Load formatting keywords from a JSON file.

        :return: Dictionary containing formatting keywords
        """
        jsonPath = os.path.join(self.projectRoot, "resources", "TAB_ApoioTitles.json")
        with open(jsonPath, 'r', encoding='utf-8') as file:
            return json.load(file)

    def analyzeData(self, excelFile: str) -> None:
        """
        Analyze data from an Excel file.

        :param excelFile: Path to the Excel file to be analyzed
        """
        logging.info(f"Starting analysis of file: {excelFile}")
        workbook = load_workbook(excelFile)
        df = pd.read_excel(excelFile, sheet_name='TAB_EvolRazSoc', header=5)
        df = df.dropna(how='all').reset_index(drop=True)
        df = df.dropna(subset=['MUNICIPIO']).reset_index(drop=True)
        df = df.fillna(0)
        df['InscEst'] = df['InscEst'].astype(str)
        
        municipios = df['MUNICIPIO'].unique()
        logging.info(f"Municipalities to be analyzed: {municipios}")
        
        for municipio in municipios:
            dfMun = df[df['MUNICIPIO'] == municipio].copy()
            self.analyzeMunicipio(dfMun, workbook, excelFile)
        
        workbook.save(excelFile)
        logging.info("Analysis completed and file saved.")

    def analyzeMunicipio(self, dfMun: pd.DataFrame, workbook, excelFile: str) -> None:
        """
        Analyze data for a specific municipality.

        :param dfMun: DataFrame containing data for the municipality
        :param workbook: Excel workbook object
        :param excelFile: Path to the Excel file to be analyzed
        """
        municipio = dfMun['MUNICIPIO'].iloc[0]
        sigMun = self.sigMunMap.get(municipio, "")
        sheetName = f"Analise{sigMun}"
        logging.info(f"Analyzing municipality: {municipio} (Abbreviation: {sigMun})")
        
        if sheetName in workbook.sheetnames:
            del workbook[sheetName]
            logging.info(f"Existing sheet {sheetName} deleted.")
        
        workbook.create_sheet(sheetName)
        sheet = workbook[sheetName]
        logging.info(f"New sheet {sheetName} created.")

        # Insert titles in the sheet
        self.insertTitles(sheet, municipio)        
        
        totalByYear = self.calculateTotalByYear(dfMun)
        totalContributors = len(dfMun)
        trendCounts, topTrendsLast, topTrendsFull = self.analyzeTrends(dfMun)
        standardDeviation = self.calculateStandardDeviation(dfMun)
        topContributors = self.getTopContributors(dfMun)
        zeroMovement = self.identifyZeroMovement(dfMun)
        
        self.updateExcel(sheet, totalByYear, totalContributors,
                          trendCounts, standardDeviation, topTrendsLast, topTrendsFull, 
                          topContributors, zeroMovement)
        
    def insertTitles(self, sheet, municipio: str):
        """
        Insert titles in the Excel sheet for a given municipality.

        :param sheet: Excel sheet where titles will be inserted
        :param municipio: Name of the municipality
        """
        logging.debug(f"Inserting titles for municipality: {municipio}")

        startTitle1 = self.config.getint('FORMATTING', 'start_title1')  
        cellTitle1 = sheet.cell(row=startTitle1, column=1, value="ESTUDO DA EVOLUÇÃO - VALOR ADICIONADO")
        cellTitle1.font = self.getFontConfiguration('title1')
        logging.debug(f"Title inserted at line {startTitle1}: ESTUDO DA EVOLUÇÃO - VALOR ADICIONADO")
        
        # Municipality title
        startTitle2 = self.config.getint('FORMATTING', 'start_title2') 
        cellTitle2 = sheet.cell(row=startTitle2, column=1, value=f"Município - {municipio}")
        cellTitle2.font = self.getFontConfiguration('title2')
        logging.debug(f"Title inserted at line {startTitle2}: Município - {municipio}")

        # Title with current date
        startTitle3 = self.config.getint('FORMATTING', 'start_title3') 
        currentDate = datetime.now().strftime("%d/%m/%Y")
        cellTitle3 = sheet.cell(row=startTitle3, column=1, value=f"Relatório Calculado em: {currentDate}")
        cellTitle3.font = self.getFontConfiguration('title3')
        logging.debug(f"Title inserted at line {startTitle3}: Relatório Calculado em: {currentDate}")

    def calculateTotalByYear(self, df: pd.DataFrame) -> pd.Series:
        """
        Calculate the total value aggregated by year.

        :param df: DataFrame containing data to be analyzed
        :return: Series containing total values by year
        """
        return df.filter(regex='^20(1[7-9]|2[0-3])').sum()

    def analyzeTrends(self, df: pd.DataFrame) -> Tuple[Dict[str, int], Dict[str, pd.DataFrame], Dict[str, pd.DataFrame]]:
        """
        Analyze trends in the data.

        :param df: DataFrame containing data to be analyzed
        :return: Tuple containing trend counts and top trends
        """
        allYears = sorted([col for col in df.columns if col.startswith('20')])
        initialYear = self.config['ANALYSIS']['InitialYear']
        lastYear = allYears[-1]
        penultimateYear = allYears[-2]
        
        minThreshold = float(self.config['ANALYSIS']['MinimumAnalysisThresholdPercentage']) / 100
        totalLastYear = df[lastYear].sum()
        minValue = totalLastYear * minThreshold
        
        dfFiltered = df[df[lastYear] >= minValue].copy()
        
        dfFiltered.loc[:, 'variationPctFull'] = self.calculateVariation(dfFiltered, initialYear, lastYear)
        dfFiltered.loc[:, 'variationPctLast'] = self.calculateVariation(dfFiltered, penultimateYear, lastYear)
        dfFiltered.loc[:, 'variationAbsFull'] = dfFiltered[lastYear] - dfFiltered[initialYear]
        dfFiltered.loc[:, 'variationAbsLast'] = dfFiltered[lastYear] - dfFiltered[penultimateYear]
        
        growthLimit = float(self.config['ANALYSIS']['SignificantPositiveVariation']) / 100
        declineLimit = float(self.config['ANALYSIS']['SignificantNegativeVariation']) / 100
        
        trendCounts = {
            'CRESCIMENTO': sum(dfFiltered['variationPctLast'] > growthLimit),
            'ESTÁVEL': sum((dfFiltered['variationPctLast'] >= 0) & (dfFiltered['variationPctLast'] <= 0.005)),
            'DECLÍNIO': sum(dfFiltered['variationPctLast'] < 0)
        }
        
        topTrendsLast = self.prepareTopTrends(dfFiltered, penultimateYear, lastYear, 'variationPctLast', 'variationAbsLast')
        topTrendsFull = self.prepareTopTrends(dfFiltered, initialYear, lastYear, 'variationPctFull', 'variationAbsFull')
        
        return trendCounts, topTrendsLast, topTrendsFull

    def calculateVariation(self, df: pd.DataFrame, startYear: str, endYear: str) -> pd.Series:
        """
        Calculate the variation percentage between two years.

        :param df: DataFrame containing data to be analyzed
        :param startYear: Starting year for the calculation
        :param endYear: Ending year for the calculation
        :return: Series containing variation percentages
        """
        return np.where(df[startYear] != 0,
                        (df[endYear] - df[startYear]) / df[startYear],
                        np.where(df[endYear] != 0, 1, 0))

    def prepareTopTrends(self, df: pd.DataFrame, startYear: str, endYear: str, varPctCol: str, varAbsCol: str) -> Dict[str, pd.DataFrame]:
        """
        Prepare top trends for growth, stability, and decline.

        :param df: DataFrame containing data to be analyzed
        :param startYear: Starting year for the calculation
        :param endYear: Ending year for the calculation
        :param varPctCol: Column name for percentage variation
        :param varAbsCol: Column name for absolute variation
        :return: Dictionary containing DataFrames for each trend
        """
        topTrends = {}
        for trend, condition in [('CRESCIMENTO', df[varPctCol] > 0.005),
                                 ('ESTÁVEL', (df[varPctCol] >= 0) & (df[varPctCol] <= 0.005)),
                                 ('DECLÍNIO', df[varPctCol] < 0)]:
            topDf = df[condition].nlargest(15, varPctCol)
            topTrends[trend] = topDf[['RazSoc', 'InscEst', startYear, endYear, varPctCol, varAbsCol]]
            topTrends[trend].columns = ['NOME / RAZÃO SOCIAL', 'InscEst', f'Valor {startYear}', f'Valor {endYear}', 'Variação %', 'Variação R$']
            topTrends[trend] = topTrends[trend].sort_values('Variação %', ascending=False)
        return topTrends

    def calculateStandardDeviation(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate the standard deviation for each entry across the selected years.

        :param df: DataFrame containing data to be analyzed
        :return: DataFrame with standard deviation, mean, and median
        """
        logging.info("Calculating standard deviation")
        initialYear = self.config['ANALYSIS']['InitialYear']
        allYears = [col for col in df.columns if col.startswith('20') and col >= initialYear]
        
        dfFiltered = df[df[allYears].min(axis=1) > 0].copy()
        
        dfFiltered['VALOR DP'] = dfFiltered[allYears].std(axis=1)
        dfFiltered['MÉDIA'] = dfFiltered[allYears].mean(axis=1)
        dfFiltered['MEDIANA'] = dfFiltered[allYears].median(axis=1)
        
        topCount = int(self.config['ANALYSIS']['StandardDeviation'])
        result = dfFiltered.nlargest(topCount, 'VALOR DP')[['RazSoc', 'InscEst', 'VALOR DP', 'MÉDIA', 'MEDIANA']]
        
        logging.debug(f"Standard deviation calculation result:\n{result.head()}")
        return result

    def getTopContributors(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Get the top contributors for the last year.

        :param df: DataFrame containing data to be analyzed
        :return: DataFrame with top contributors
        """
        lastYear = df.filter(regex='^20').columns[-1]
        topCount = int(self.config['ANALYSIS']['TopContributors'])
        contributors = df.nlargest(topCount, lastYear)[['RazSoc', 'InscEst', lastYear]].reset_index(drop=True)
        contributors.columns = ['NOME / RAZÃO SOCIAL', 'InscEst', 'Contribuição']
        return contributors

    def identifyZeroMovement(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Identify contributors with zero movement over the years.

        :param df: DataFrame containing data to be analyzed
        :return: DataFrame with contributors with zero movement
        """
        zeroMovement = df[df.filter(regex='^20(1[7-9]|2[0-3])').sum(axis=1) == 0][['RazSoc', 'InscEst']].reset_index(drop=True)
        zeroMovement.columns = ['NOME / RAZÃO SOCIAL', 'InscEst']
        return zeroMovement

    def updateExcel(self, sheet, *args):
        """
        Update the Excel sheet with calculated data.

        :param sheet: Excel sheet to be updated
        :param args: Arguments containing data to be inserted
        """
        logging.info("Updating Excel sheet")
        self.setupStyles(sheet.parent)
        
        row = 7
        titles = self.getTitles(args[0])
        
        for title, arg in zip(titles, args):
            logging.debug(f"Processing: {title}")
            
            # Write the title
            cellTitle = sheet.cell(row=row, column=1, value=title)
            cellTitle.font = self.getFontConfiguration('title3')
            cellTitle.alignment = Alignment(horizontal='left', vertical='center')
            row += 1

            # Process the data
            if title == "TOTAL DE CONTRIBUINTES":
                cellValue = sheet.cell(row=row-1, column=2, value=int(arg))
                cellValue.font = self.getFontConfiguration('normal')
                cellValue.number_format = '0'  # Integer number format
            elif isinstance(arg, pd.DataFrame):
                row = self.processDataFrame(sheet, row, title, arg)
            elif isinstance(arg, pd.Series):
                row = self.processSeries(sheet, row, title, arg)
            elif isinstance(arg, dict):
                row = self.processDict(sheet, row, title, arg)
            elif isinstance(arg, (int, float)):
                cellValue = sheet.cell(row=row, column=1, value=arg)
                cellValue.font = self.getFontConfiguration('normal')
                self.applyStyle(cellValue, title, title)
                row += 1
            else:
                logging.warning(f"Unrecognized data type for {title}: {type(arg)}")
            
            row += int(self.config['ANALYSIS']['BlockSpacing'])
        
        self.adjustColumnWidths(sheet)
        logging.info("Excel sheet update completed")

    def setupStyles(self, workbook):
        """
        Setup named styles for accounting and percentage formats in the workbook.

        :param workbook: Excel workbook where styles will be added
        """
        if "accounting_style" not in workbook.named_styles:
            accountingStyle = NamedStyle(name="accounting_style")
            accountingStyle.number_format = self.config['FORMATTING']['accounting_format']
            workbook.add_named_style(accountingStyle)

        if "percent_style" not in workbook.named_styles:
            percentStyle = NamedStyle(name="percent_style")
            percentStyle.number_format = self.config['FORMATTING']['percent_format']
            workbook.add_named_style(percentStyle)

    def getFontConfiguration(self, styleType: str = 'normal') -> Font:
        """
        Get font configuration based on the style type.

        :param styleType: Style type ('normal', 'title1', 'title2', 'title3')
        :return: Font object with specified configuration
        """
        fontConfig = {
            'normal': {
                'font': self.config.get('FORMATTING', 'font_normal'),
                'size': self.config.getint('FORMATTING', 'font_size_normal'),
                'bold': self.config.getboolean('FORMATTING', 'font_style_normal'),
                'italic': False
            },
            'title1': {
                'font': self.config.get('FORMATTING', 'font_title'),
                'size': self.config.getint('FORMATTING', 'font_size_title1'),
                'bold': self.config.getboolean('FORMATTING', 'font_style_bold'),
                'italic': False
            },
            'title2': {
                'font': self.config.get('FORMATTING', 'font_title'),
                'size': self.config.getint('FORMATTING', 'font_size_title2'),
                'bold': self.config.getboolean('FORMATTING', 'font_style_bold'),
                'italic': False
            },
            'title3': {
                'font': self.config.get('FORMATTING', 'font_title'),
                'size': self.config.getint('FORMATTING', 'font_size_title3'),
                'bold': self.config.getboolean('FORMATTING', 'font_style_bold'),
                'italic': False
            }
        }

        config = fontConfig[styleType]
        return Font(
            name=config['font'],
            size=config['size'],
            bold=config['bold'],
            italic=config['italic']
        )        

    def getTitles(self, data):
        """
        Get titles for the Excel sheet sections based on the data.

        :param data: Data used to determine the titles
        :return: List of titles for the sections
        """
        allYears = sorted([int(col) for col in data.index if col.isdigit()])
        firstYear = self.config['ANALYSIS']['InitialYear']
        lastYear, penultimateYear = str(allYears[-1]), str(allYears[-2])
        return [
            "VALOR TOTAL AGREGADO POR ANO",
            "TOTAL DE CONTRIBUINTES",
            "CONTAGEM DE TENDÊNCIAS",
            f"DESVIO PADRÃO {firstYear} - {lastYear}",  # Modificado
            f"TENDÊNCIA {penultimateYear} / {lastYear}",
            f"TENDÊNCIA {firstYear} / {lastYear}",
            "PRINCIPAIS CONTRIBUINTES",
            f"CONTRIBUINTES SEM MOVIMENTAÇÃO DE {firstYear} À {lastYear}"
        ]

    def processSection(self, sheet, row, title, arg):
        """
        Process a section of data for the Excel sheet.

        :param sheet: Excel sheet to be updated
        :param row: Starting row for the section
        :param title: Title of the section
        :param arg: Data to be processed
        :return: Updated row number after processing
        """
        logging.debug(f"Processing section: {title}")
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = self.getFontConfiguration('title2')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        if isinstance(arg, pd.DataFrame):
            row = self.processDataFrame(sheet, row, title, arg)
        elif isinstance(arg, pd.Series):
            row = self.processSeries(sheet, row, title, arg)
        elif isinstance(arg, (int, float)):
            row = self.processSingleValue(sheet, row, title, arg)
        elif isinstance(arg, dict):
            row = self.processDict(sheet, row, title, arg)

        return row + int(self.config['ANALYSIS']['BlockSpacing'])

    def processDataFrame(self, sheet, row, title, df):
        """
        Process a DataFrame and insert it into the Excel sheet.

        :param sheet: Excel sheet to be updated
        :param row: Starting row for the DataFrame
        :param title: Title of the section
        :param df: DataFrame to be inserted
        :return: Updated row number after processing
        """
        logging.debug(f"Processing DataFrame for {title}")
        if df.empty:
            cell = sheet.cell(row=row, column=1, value="*** NENHUMA EMPRESA ATENDEU ESTE QUESITO ***")
            cell.font = self.getFontConfiguration('normal')
            return row + 1

        for col, columnName in enumerate(df.columns, start=1):
            cell = sheet.cell(row=row, column=col, value=columnName.upper())
            cell.font = self.getFontConfiguration('title3')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        for _, data in df.iterrows():
            for col, value in enumerate(data, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.font = self.getFontConfiguration('normal')
                self.applyStyle(cell, df.columns[col-1], title)
            row += 1

        # Add explanatory note for standard deviation
        if "DESVIO PADRÃO" in title:
            row += 1
            note = "Nota Explicativa: O desvio padrão indica a variabilidade das contribuições ao longo do período analisado."
            cell = sheet.cell(row=row, column=1, value=note)
            cell.font = self.getFontConfiguration('title3')
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(df.columns))

        return row

    def processSeries(self, sheet, row, title, series):
        """
        Process a Series and insert it into the Excel sheet.

        :param sheet: Excel sheet to be updated
        :param row: Starting row for the Series
        :param title: Title of the section
        :param series: Series to be inserted
        :return: Updated row number after processing
        """
        logging.debug(f"Processing Series for {title}")
        for index, value in series.items():
            sheet.cell(row=row, column=1, value=index).font = self.getFontConfiguration('title3')
            cell = sheet.cell(row=row, column=2, value=value)
            cell.font = self.getFontConfiguration('normal')

            self.applyStyle(cell, index, title)
            row += 1
        return row

    def processSingleValue(self, sheet, row, title, value):
        """
        Process a single value and insert it into the Excel sheet.

        :param sheet: Excel sheet to be updated
        :param row: Starting row for the value
        :param title: Title of the section
        :param value: Value to be inserted
        :return: Updated row number after processing
        """
        logging.debug(f"Processing single value for {title}")
        cellTitle = sheet.cell(row=row, column=1, value=title)
        cellTitle.font = self.getFontConfiguration('title3')
        
        if title == "TOTAL DE CONTRIBUINTES":
            cellValue = sheet.cell(row=row, column=2, value=int(value))  # Convert to integer
            cellValue.font = self.getFontConfiguration('normal')
            cellValue.number_format = '0'  # Integer number format
        else:
            cellValue = sheet.cell(row=row, column=1, value=value)
            cellValue.font = self.getFontConfiguration('normal')
            self.applyStyle(cellValue, title, title)
        
        return row + 1

    def processDict(self, sheet, row, title, arg):
        """
        Process a dictionary and insert its contents into the Excel sheet.

        :param sheet: Excel sheet to be updated
        :param row: Starting row for the dictionary
        :param title: Title of the section
        :param arg: Dictionary to be processed
        :return: Updated row number after processing
        """
        logging.debug(f"Processing dictionary for {title}")
        if title.startswith("TENDÊNCIA"):
            for trend, df in arg.items():
                if trend in ["ESTÁVEL", "DECLÍNIO"]:
                    row += 2
                trendCell = sheet.cell(row=row, column=1, value=trend.upper())
                trendCell.font = self.getFontConfiguration('title3')
                trendCell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                if isinstance(df, pd.DataFrame):
                    row = self.processDataFrame(sheet, row, trend, df)
            return row + 2
        else:
            for key, value in arg.items():
                sheet.cell(row=row, column=1, value=key).font = self.getFontConfiguration('title3')
                cell = sheet.cell(row=row, column=2, value=value)
                cell.font = self.getFontConfiguration('normal')
                self.applyStyle(cell, key, title)
                row += 1
            return row

    def applyStyle(self, cell, columnName, title):
        """
        Apply the appropriate style to a cell based on its content.

        :param cell: Cell to which the style will be applied
        :param columnName: Name of the column containing the cell
        :param title: Title of the section containing the cell
        """
        columnName = self.normalizeString(columnName)
        title = self.normalizeString(title)
        
        logging.debug(f"Applying style for: {columnName} (title: {title})")
        
        for keyword in self.formatKeywords['monetary_keywords']:
            if self.normalizeString(keyword) in columnName or self.normalizeString(keyword) in title:
                logging.debug(f"Monetary style applied for: {columnName}")
                cell.style = 'accounting_style'
                cell.font = self.getFontConfiguration('normal')
                return
        
        for keyword in self.formatKeywords['percentage_keywords']:
            if self.normalizeString(keyword) in columnName or self.normalizeString(keyword) in title:
                logging.debug(f"Percentage style applied for: {columnName}")
                cell.style = 'percent_style'
                cell.font = self.getFontConfiguration('normal')
                return
        
        logging.debug(f"No specific style applied for: {columnName}")

    def normalizeString(self, s):
        """
        Normalize a string by converting it to lowercase and stripping whitespace.

        :param s: String to be normalized
        :return: Normalized string
        """
        return ' '.join(s.lower().strip().split())

    def adjustColumnWidths(self, sheet):
        """
        Adjust the width of columns and set the page layout to landscape in the Excel sheet.

        :param sheet: Excel sheet where adjustments will be made
        """
        for colLetter in ['B', 'C', 'D', 'E', 'F', 'G']:
            sheet.column_dimensions[colLetter].width = 15

        sheet.column_dimensions['A'].width = 51
        for cell in sheet['A']:
            cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)

        sheet.page_setup.orientation = 'landscape'
        sheet.page_margins = PageMargins(left=0.3, right=0, top=0.3, bottom=0, header=0.1, footer=0)

        maxRow = sheet.max_row
        sheet.print_area = f'A1:F{maxRow}'
        sheet.print_title_rows = '1:4'

        sheet.oddHeader.center.text = "&P / &N"
        sheet.oddHeader.center.size = 8  # Font size
        sheet.oddHeader.center.font = "Arial"

def main():
    """
    Main function to setup the DataAnalyzer and start the data analysis process.
    """
    projectRoot = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config = configparser.ConfigParser()
    config.read(os.path.join(projectRoot, "resources", "Config.ini"))
    analyzer = DataAnalyzer(projectRoot, config)
    excelFile = os.path.join(projectRoot, 
                             config['DEFAULT']['OutputDirectory'], 
                             config['DEFAULT']['OutputFileName'])
    analyzer.analyzeData(excelFile)

if __name__ == "__main__":
    main()    